import pymysql
import openpyxl
from openpyxl.chart import BarChart, Reference, Series
import string


class AdminDB:
    _NAME_DB = 'elen_db'
    _NAME_TABLE_USERS = 'users'
    _NAME_TABLE_LOG = 'log'
    _NAME_TABLE_HISTORY = 'history'
    _NAME_TABLE_TARIFFS = 'tariffs'
    _HOST_DB = 'localhost'
    _USER_DB = 'root'
    _PASSWORD_DB = '123456'

    STARTING_TARIFFS = 'starting tariffs', '2015-09-01', 100, 300, 0.5, 1.0, 2.0

    def __init__(self):
        try:
            self.connect_db()
        except pymysql.err.InternalError:
            try:
                self.create_db()
                self.use_db()
                self.create_table(self._NAME_TABLE_USERS)
                self.create_table('{}'.format(self._NAME_TABLE_TARIFFS),
                                  'limit_tariff_1', 'limit_tariff_2', 'tariff_1', 'tariff_2', 'tariff_3'
                                  )
                self.write_into('{}'.format(self._NAME_TABLE_TARIFFS),
                                self.STARTING_TARIFFS)
            except:
                raise ConnectionError

    def create_db(self):
        print('create_db', self._NAME_DB)
        self.conn = pymysql.connect(host=self._HOST_DB,
                                    user=self._USER_DB,
                                    password=self._PASSWORD_DB)
        self.cur = self.conn.cursor()
        self.cur.execute('CREATE DATABASE {}'.format(self._NAME_DB))

    def use_db(self):
        print('use_db', self._NAME_DB)
        self.cur.execute('USE {}'.format(self._NAME_DB))

    def connect_db(self):
        print('connect_db', self._NAME_DB)
        self.conn = pymysql.connect(host=self._HOST_DB,
                                    user=self._USER_DB,
                                    password=self._PASSWORD_DB,
                                    db=self._NAME_DB,
                                    use_unicode=True,
                                    charset="utf8"
                                    )
        self.cur = self.conn.cursor()

    def create_table(self, name_table, *args):
        print('create_table', name_table)
        if name_table == self._NAME_TABLE_USERS:
            self.cur.execute('CREATE TABLE {}('
                             'id integer AUTO_INCREMENT PRIMARY KEY, '
                             'user varchar(20) NOT NULL, '
                             'password varchar(40) NOT NULL)'
                             .format(name_table)
                             )
        elif name_table == self._NAME_TABLE_TARIFFS:
            self.cur.execute('CREATE TABLE {}('
                             'user varchar(20) NOT NULL PRIMARY KEY, '
                             'date_now DATE NOT NULL, '
                             .format(name_table)
                             +
                             str([arg + ' varchar(20) NOT NULL' for arg in args])[1:-1].replace("'", "")
                             +
                             ')'
                             )
        elif self._NAME_TABLE_LOG in name_table:
            self.cur.execute('CREATE TABLE {}('
                             'id MEDIUMINT NOT NULL AUTO_INCREMENT PRIMARY KEY, '
                             'date_introduction_tariffs DATE NOT NULL, '
                             'date_now DATE NOT NULL, '
                             .format(name_table)
                             +
                             str([arg + ' varchar(20) NOT NULL' for arg in args])[1:-1].replace("'", "")
                             +
                             ')'
                             )
        elif self._NAME_TABLE_HISTORY in name_table:
            self.cur.execute('CREATE TABLE {}('
                             #'id MEDIUMINT NOT NULL AUTO_INCREMENT, '
                             'date_introduction_tariffs DATE NOT NULL PRIMARY KEY, '
                             'date_now DATE NOT NULL, '
                             .format(name_table)
                             +
                             str([arg + ' DECIMAL(10,2) NOT NULL' for arg in args])[1:-1].replace("'", "")
                             +
                             ')'
                             )

    def create_user(self, user_name, user_password):
        print('create_user', user_name)
        self.cur.execute('INSERT INTO users(user, password) VALUES(%s, %s)', (user_name, user_password))
        self.conn.commit()

    def check_user(self, user_name):
        print('check_user', user_name)
        self.cur.execute('SELECT user FROM users WHERE user="{}"'.format(user_name))
        return self.cur.fetchall()

    def get_columns_list_without_id(self, table):
        print('get_columns_list_without_id', table)
        self.cur.execute('SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS '
                         'WHERE TABLE_NAME="{}" AND COLUMN_NAME NOT IN ("id")'.format(table)
                         )
        columns_list_without_id = '(' + str(self.cur.fetchall())[1:].replace("('", '').replace("',)", '')
        print(columns_list_without_id)
        return columns_list_without_id

    def write_into(self, table, info):
        print('write_into', table, info)
        self.cur.execute('INSERT INTO {} {} VALUES {}'
                         .format(table, self.get_columns_list_without_id(table), info)
                         )
        self.conn.commit()

    def del_from(self, table, key):
        pass

    def update(self, table, info):
        print('update', table, info)
        self.cur.execute('REPLACE INTO {} VALUES {}'.format(table, info))
        self.conn.commit()

    def check_password(self, user_name, user_password):
        print('check_password', user_name)
        self.cur.execute('SELECT password '
                         'FROM users '
                         'WHERE user="{}"'
                         .format(user_name)
                         )
        return self.cur.fetchone()[0]

    def read_saved_tariffs(self, user_name):
        print('read_saved_tariffs', user_name)
        self.cur.execute('SELECT * '
                         'FROM tariffs '
                         'WHERE user="{}"'
                         .format(user_name)
                         )
        tariffs_from_db = self.cur.fetchone()[1:]
        return tariffs_from_db

    def get_sorted(self, table, column_name):
        print('sort_by', table, column_name)
        self.cur.execute('SELECT * FROM {} ORDER BY {}'
                         .format(table, column_name)
                         )
        print(self.cur.fetchall())

    def read_all_from(self, table):
        print('read_all_from', table)
        self.cur.execute('SELECT * FROM {}'.format(table))
        data_from_db = self.cur.fetchall()
        return data_from_db

    def get_count_rows(self, table):
        self.cur.execute('SELECT COUNT(*) FROM {}'.format(table))
        count_rows_from_db = self.cur.fetchone()[0]
        return count_rows_from_db

    def get_number_for_letter(self, letter):
        return string.ascii_uppercase.index(letter)

    def get_letter_for_number(self, number):
        return list(string.ascii_uppercase)[number]

    def set_head_table_in_excel(self, ws, table, start_row_in_excel, start_column_in_excel):
        columns_list_from_db = self.get_columns_list_without_id(table)[1:-1].split(', ')
        for number_column, name_column_in_db in enumerate(columns_list_from_db,
                                                          self.get_number_for_letter(start_column_in_excel)):
            ws['{}{}'.format(self.get_letter_for_number(number_column), start_row_in_excel)] = name_column_in_db
        ws.auto_filter.ref = '{}{}:{}{}'.format(start_column_in_excel,
                                                start_row_in_excel,
                                                self.get_letter_for_number(
                                                    self.get_number_for_letter(start_column_in_excel) +
                                                    len(columns_list_from_db)-1),
                                                start_row_in_excel + self.get_count_rows(table)
                                                )

    def set_body_table_in_excel(self, ws, data_from_db, start_row_in_excel, start_column_in_excel):
        for number_row, row in enumerate(data_from_db, start_row_in_excel + 1):
            for number_column, value in enumerate(row, self.get_number_for_letter(start_column_in_excel)):
                ws['{}{}'.format(self.get_letter_for_number(number_column), number_row)] = value

    def build_graph_in_excel(self, ws, table, start_row_in_excel, start_column_in_excel):
        columns_list_from_db = self.get_columns_list_without_id(table)[1:-1].split(', ')
        count_rows_in_table_db = self.get_count_rows(table)+1
        start_column_number = self.get_number_for_letter(start_column_in_excel) + 1

        chart = BarChart()
        chart.title = '{}'.format(table)
        chart.style = 10
        #chart.x_axis.title = 'TEst'
        chart.y_axis.title = 'Percentage'

        cats = Reference(ws,
                         min_col=start_column_number,
                         min_row=start_row_in_excel + 1,
                         max_row='{}'.format(count_rows_in_table_db)
                         )
        data1 = Reference(ws,
                          min_col=10,
                          min_row=1,
                          max_row='{}'.format(count_rows_in_table_db)
                          )
        data2 = Reference(ws, min_col=14, min_row=1, max_row='{}'.format(count_rows_in_table_db))

        chart.add_data(data1, titles_from_data=True)
        chart.add_data(data2, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, 'A15')


    def export_history_to_excel(self, table, date):
        print('export_in_excel_db', table)
        data_from_db = self.read_all_from(table)
        print('data_from_db', data_from_db)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '{}'.format(table)
        start_row_in_excel = 1
        start_column_in_excel = 'A'
        self.set_head_table_in_excel(ws, table, start_row_in_excel, start_column_in_excel)
        self.set_body_table_in_excel(ws, data_from_db, start_row_in_excel, start_column_in_excel)
        self.build_graph_in_excel(ws, table, start_row_in_excel, start_column_in_excel)
        wb.save('elen_export_{}_{}.xlsx'.format(table, date))



    def close_connection(self):
        print('close_connection')
        self.conn.close()